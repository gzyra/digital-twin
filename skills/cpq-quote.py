import os
import time
import openpyxl
import configparser
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select

from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options as ChromeOptions
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from webdriver_manager.firefox import GeckoDriverManager




def print_all_elements_from_driver(driver):

    driver_elements = driver.find_elements(By.XPATH, '//*')

    print(50*'*')
    print('Trying to print all By.XPATH, "//*" elements from driver')
    print(50*'*')

    i = 0
    for element in driver_elements:
        print()
        print('i = {}'.format(i))
        print('Element id: {}'.format(element.id))
        print('Element text: {}'.format(element.text))
        print('Element tag_name: {}'.format(element.tag_name))
        print('Element value_of_css_property: {}'.format(element.value_of_css_property))
        print('Element is_displayed: {}'.format(element.is_displayed))
        print('Element is_selected: {}'.format(element.is_selected))
        print('Element is_enabled: {}'.format(element.is_enabled))
        print(20*'-')
        print('Values of get_attibute:')
        print('Element id: {}'.format(element.get_attribute('id')))
        print('Element text: {}'.format(element.get_attribute('text')))
        print('Element class: {}'.format(element.get_attribute('class')))
        print('Element name: {}'.format(element.get_attribute('name')))
        print('Element type: {}'.format(element.get_attribute('type')))
        print('Element href: {}'.format(element.get_attribute('href')))
        print('Element tag: {}'.format(element.get_attribute('tag')))
        print('Element label: {}'.format(element.get_attribute('label')))
        print('Element index: {}'.format(element.get_attribute('index')))
        print('Element item: {}'.format(element.get_attribute('item')))
        print('Element slot: {}'.format(element.get_attribute('slot')))
        print('Element role: {}'.format(element.get_attribute('role')))
        print('Element filter-key: {}'.format(element.get_attribute('filter-key')))
        print('Element kdfapp: {}'.format(element.get_attribute('kdfapp')))
        print('Element kdfpage: {}'.format(element.get_attribute('kdfpage')))
        print('Element kdfid: {}'.format(element.get_attribute('kdfid')))

        i = i + 1

    # print(dir(element))


def print_driver_page_source(driver):

    print(50*'*')
    print('Printing driver source')
    print(50*'*')
    print(driver.page_source)


def print_driver_windows(driver):

    print(50*'*')
    print('Printing driver windows handler')
    print(50*'*')

    for window_handle in driver.window_handles:
        driver.switch_to.window(window_handle)
        print('Window ID (handle): {}, Window Title: {}'.format(window_handle, driver.title))
        

def find_all_iframes(driver):


    # iframes = driver.find_elements_by_xpath("//iframe")
    # for index, iframe in enumerate(iframes):

    #     # Your sweet business logic applied to iframe goes here.
    #     driver.switch_to.frame(index)
    #     find_all_iframes(driver)
    #     driver.switch_to.parent_frame()

    print(50*'*')
    print('Printing iframes from driver')
    print(50*'*')

    iframes = driver.find_elements(By.XPATH, '//iframe')
    
    for iframe in iframes:
        print(iframe)
      
# END OF TOOL KIT










class Opportunity():

    def __init__(self):
        self.ngsc_theater = 'EMEA'
        self.ngsc_region = 'EMEA UKI'
        self.delivery_region = 'UK'
        self.quote_name = ''
        self.customer_name = ''
        self.line_of_business = 'ENTERPRISE'
        self.sales_name = 'gzyra'
        self.DID = '12345678'
        self.PID = ''
        self.project_description = ''
        self.reviewer_name = ''

    def __str__(self):
        tw = 25
        return  '{:<{cw}}: {:<}\n'.format('NGSC Theater', self.ngsc_theater, cw=tw) + \
                '{:<{cw}}: {:<}\n'.format('NGSC Region', self.ngsc_region, cw=tw) + \
                '{:<{cw}}: {:<}\n'.format('Delivery Region', self.delivery_region, cw=tw) + \
                '{:<{cw}}: {:<}\n'.format('Quote Name', self.quote_name, cw=tw) + \
                '{:<{cw}}: {:<}\n'.format('Customer Name', self.customer_name, cw=tw) + \
                '{:<{cw}}: {:<}\n'.format('Line of Business', self.line_of_business, cw=tw) + \
                '{:<{cw}}: {:<}\n'.format('Sales Name', self.sales_name, cw=tw) + \
                '{:<{cw}}: {:<}\n'.format('DID', self.DID, cw=tw) + \
                '{:<{cw}}: {:<}\n'.format('PID', self.PID, cw=tw) + \
                '{:<{cw}}: {:<}\n'.format('Project Description', self.project_description, cw=tw) + \
                '{:<{cw}}: {:<}\n'.format('Quote Reviewer Name (DM)', self.reviewer_name, cw=tw)

class Service_Task():
    def __init__(self):
        self.ms = ''
        self.ms_desc = ''
        self.task_desc = ''
        self.ATO_mapping = ''
        self.SKU = ''
        self.primary_resource_h = ''
        self.primary_resource_g = ''
        self.project_manager_h = ''
        self.project_manager_g = ''
        self.trips_nr = ''
        self.nights_nr = ''
        self.total_TE_cost = ''

    def __str__(self):
        return  'MS: {} | '.format(self.ms) + \
                '{} | '.format(self.ms_desc) + \
                '{} | '.format(self.task_desc) + \
                '{} | '.format(self.ATO_mapping) + \
                '{} | '.format(self.SKU) + \
                '{} | '.format(self.primary_resource_h) + \
                '{} | '.format(self.primary_resource_g) + \
                '{} | '.format(self.project_manager_h) + \
                '{} | '.format(self.project_manager_g) + \
                '{} | '.format(self.trips_nr) + \
                '{} | '.format(self.nights_nr) + \
                '{} | '.format(self.total_TE_cost)


def print_text(test_to_print):
    
    print()
    print(60 * '-')
    print(test_to_print)
    print(60 * '-')


def setup_web_driver(browser):
    
    if browser == 'Firefox':
        firefox_options = FirefoxOptions()
        firefox_options.set_preference('dom.webnotifications.enabled', False)
        driver = webdriver.Firefox(service=FirefoxService(GeckoDriverManager().install()), options=firefox_options)
    

    elif browser == 'Chrome':
        chrome_options = ChromeOptions()
        chrome_options.add_argument('--disable-search-engine-choice-screen')
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)

    else:
        print('Incorrect Browser specification. Firefox has been set as brower for webdriver.')
        driver = webdriver.Firefox(service=FirefoxService(GeckoDriverManager().install()))

    driver.maximize_window()
        
    return driver


def read_script_input_data_file(input_file_name, srcript_input):

    config_file_name = input_file_name

    parser = configparser.ConfigParser()
    
    # --------------------------------------------------------------------------------------------------------------
    # Reading starting parameters from config_file_name file
    # --------------------------------------------------------------------------------------------------------------
    if os.path.exists(config_file_name):

        with open(config_file_name) as f:
            parser.read_file(f)

        # --------------------------------------------------------------------------------------------------------------
        # General
        # --------------------------------------------------------------------------------------------------------------  
        srcript_input['open_mode'] = parser['General']['open_mode']
        srcript_input['username'] = parser['General']['username']
        srcript_input['filename'] = parser['General']['filename']
        srcript_input['max_MS_number'] = parser['General']['maximum_number_of_milestones_for_import']

        # --------------------------------------------------------------------------------------------------------------
        # Quote Info
        # --------------------------------------------------------------------------------------------------------------
        srcript_input['opportunuty'].DID = parser['Quote Info']['cpq_quote_DID']
        srcript_input['opportunuty'].quote_name = parser['Quote Info']['cpq_quote_name']
        srcript_input['opportunuty'].ngsc_theater = parser['Quote Info']['cpq_quote_ngsc_theater']
        srcript_input['opportunuty'].ngsc_region = parser['Quote Info']['cpq_quote_ngsc_region']
        srcript_input['opportunuty'].delivery_region = parser['Quote Info']['cpq_quote_region']
        srcript_input['opportunuty'].line_of_business = parser['Quote Info']['cpq_quote_line_of_business']
        srcript_input['opportunuty'].sales_name = parser['Quote Info']['cpq_quote_sales_name']
        srcript_input['opportunuty'].reviewer_name = parser['Quote Info']['cpq_quote_reviewer_name']

    else:
        print('Missing file {} with input data which scrip needs to run.'.format(config_file_name))

    return srcript_input


def read_opp_data_from_estimates_file(filename,opp):

    '''
    Procedure to read opportunity data from LoE Buider file and return task list.
    '''

    wb = openpyxl.load_workbook(filename, data_only=True)
    s_ws = wb['LoE Builder']

    if s_ws.cell(row=2, column=3).value:
        opp.DID = s_ws.cell(row=2, column=3).value

    if s_ws.cell(row=1, column=4).value:
        opp.customer_name = s_ws.cell(row=1, column=4).value

    return opp


def read_task_data_from_estimates_file(filename,max_MS_number):

    '''
    Procedure to read all task data from LoE Buider file and return task list.
    '''

    tasks_list = []

    wb = openpyxl.load_workbook(filename, data_only=True)
    s_ws = wb['LoE Builder']
    tasks_start_row = 21
    
    for i in range (0, max_MS_number):
        ms_desciption_test = s_ws.cell(row=tasks_start_row+i*10, column=3).value
        if '<<' not in ms_desciption_test:
            for j in range (2, 10):
                if s_ws.cell(row=tasks_start_row+i*10+j, column=3).value:
                    task = Service_Task()
                    task.ms_desc = s_ws.cell(row=tasks_start_row+i*10, column=3).value
                    
                    task.ms = s_ws.cell(row=tasks_start_row+i*10+j, column=2).value
                    task.task_desc = s_ws.cell(row=tasks_start_row+i*10+j, column=3).value

                    task.ATO_mapping = s_ws.cell(row=tasks_start_row+i*10+j, column=4).value
                    task.SKU = s_ws.cell(row=tasks_start_row+i*10+j, column=5).value
                    task.primary_resource_h = s_ws.cell(row=tasks_start_row+i*10+j, column=7).value
                    task.primary_resource_g = s_ws.cell(row=tasks_start_row+i*10+j, column=11).value
                    task.project_manager_h = s_ws.cell(row=tasks_start_row+i*10+j, column=15).value
                    task.project_manager_g = s_ws.cell(row=tasks_start_row+i*10+j, column=16).value
                    task.trips_nr = s_ws.cell(row=tasks_start_row+i*10+j, column=17).value
                    task.nights_nr = s_ws.cell(row=tasks_start_row+i*10+j, column=18).value
                    task.total_TE_cost = s_ws.cell(row=tasks_start_row+i*10+j, column=31).value
                    tasks_list.append(task)

    return tasks_list


def correct_task_data(tasks_list):

    for e in tasks_list:

        e.ms_desc = e.ms_desc.replace('(', '- ')
        e.ms_desc = e.ms_desc.replace(')', '')
        e.ms_desc = e.ms_desc.replace('/', '-')
        if len(e.task_desc) > 60:
            e.task_desc = e.task_desc[:59]

        e.task_desc = e.task_desc.replace('(', '- ')
        e.task_desc = e.task_desc.replace(')', '')
        e.task_desc = e.task_desc.replace('/', '-')
        if len(e.task_desc) > 60:
            e.task_desc = e.task_desc[:59]

    return tasks_list


def passing_thru_SSO(driver, username, timeout_timer):

    # Completing infomation on first Log in page
    WebDriverWait(driver, timeout_timer).until(EC.presence_of_element_located((By.XPATH, '//input[@name="identifier"]'))).send_keys(username)
    # Pressing Next Button
    driver.find_element(By.XPATH, '//input[@type="submit" and @value="Next"]').click()
    # Moving on thru next screen
    WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '//input[@type="email"]'))).send_keys(username)
    WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.c--primary'))).click()

    # Making sure that we are on "Yes, this is my device"
    WebDriverWait(driver, timeout_timer).until(EC.presence_of_element_located((By.XPATH, '//*[text()="Yes, this is my device"]'))).click()
    
    return


def go_to_opportunity_tab(driver, opp, timeout_timer, verbose = True):

    # Entering DID number into Search Field in SalesForce
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '//button[@type="button" and @class="slds-button slds-button_neutral search-button slds-truncate"]'))).click()
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div[2]/div/div/div[1]/div/div[1]/lightning-input/lightning-primitive-input-simple/div/div/input'))).send_keys(opp.DID + Keys.RETURN)
    except Exception as e:
        if verbose:
            print('Failed with search for opportunity details for DID: {}. Error message: {}'.format(opp.DID, e.msg))
        else:
            print('Failed with search for opportunity details for DID: {}.'.format(opp.DID))

    # Selecting TAB with Search Results
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, opp.DID))).click()
    except Exception as e:
        if verbose:
            print('Failed when trying to switch search results for DID: {}. Error message: {}'.format(opp.DID, e.msg))
        else:
            print('Failed when trying to switch search results for DID: {}.'.format(opp.DID))

    # Navigating to opportunity tab
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section/div/div/section/div/div[2]/div/div/div/div[2]/div/div/div/div[3]/div/div/div/div/div[2]/div[1]/div[2]/div/div/div/div[2]/div[2]/div[1]/div/div/table/tbody/tr/th/span/a'))).click()
        # WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[class*="outputLookupLink"]'))).click()
    except Exception as e:
        if verbose:
            print('Failed when trying to open opportunity tab for DID: {}. Error message: {}'.format(opp.DID, e.msg))
        else:
            print('Failed when trying to open opportunity tab for DID: {}.'.format(opp.DID))

    return


def create_new_quote_in_CCW(driver, opp, timeout_timer, verbose = True):

    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.NAME, 'Opportunity.Quote_Technology_Service')))
        new_quote_button = driver.find_element(By.NAME, 'Opportunity.Quote_Technology_Service')
        new_quote_button.click()
    except Exception as e:
        if verbose:
            print('Failed in clicking new_quote_button. Error: {}'.format(e.msg)) 
        else:
            print('Failed in clicking new_quote_button.') 

    try:
        time.sleep(8)        
        WebDriverWait(driver, timeout_timer).until(EC.text_to_be_present_in_element((By.CSS_SELECTOR, '.slds-dropdown__item:nth-child(1) .slds-truncate'), 'Create a Quote'))
        create_a_quote_selection = driver.find_element(By.LINK_TEXT, 'Create a Quote')
        ActionChains(driver).move_to_element(create_a_quote_selection).click().perform()
    except Exception as e:
        if verbose:
            print('Failed in clicking quote_option. Time sleep {}.  Error: {}'.format(time_sleep, e.msg))
        else:
            print('Failed in clicking quote_option. Time sleep {}.'.format(time_sleep))

    # New quote creation
    # Store the ID of the original window
    saleforce_start_window = driver.current_window_handle
    
    # Wait for the new window or tab
    WebDriverWait(driver, timeout_timer).until(EC.number_of_windows_to_be(2))

    # Loop through until we find a new window handle
    for window_handle in driver.window_handles:
        if window_handle != saleforce_start_window:
            driver.switch_to.window(window_handle)
            break

    try:
        single_quote = WebDriverWait(driver, timeout_timer).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'span.slds-radio:nth-child(2) > label:nth-child(2) > b:nth-child(2) > span:nth-child(1)')))
        single_quote.click()        
    except Exception as e:
        if verbose:
            print('Failed in clicking single_quote. Error: {}'.format(e.msg))
        else:
            print('Failed in clicking single_quote.')

    try:
        next_create_quote = WebDriverWait(driver, timeout_timer).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div.slds-button:nth-child(2)')))
        next_create_quote.click()
    except Exception as e:
        if verbose:
            print('Failed in clicking next_create_quote. Error: {}'.format(e.msg)) 
        else:
            print('Failed in clicking next_create_quote.')

    #------------------------------------------------------------------------------------------
    print_text('Switching to and handling CCW page for new quote')
    #------------------------------------------------------------------------------------------
    # Switching to and handling CCW page for new quote

    time_sleep = 10
    time.sleep(time_sleep)

    # Loop through until we find a new window handle
    for window_handle in driver.window_handles:
        if window_handle != saleforce_start_window:
            driver.switch_to.window(window_handle)
            break
    
    # Setting up value of Quote Name on CCW page.
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '//input[@kdfapp="Quoting" and @kdfpage="CreateQuote" and @kdfid="quoteName"]'))).clear()
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '//input[@kdfapp="Quoting" and @kdfpage="CreateQuote" and @kdfid="quoteName"]'))).send_keys(opp.quote_name)
    except Exception as e:
        if verbose:
            print('Failed during setting up new quote name on CCW page. Error: {}'.format(e)) 
        else:
            print('Failed during setting up new quote name on CCW page.')

    # Setting up value of Intended Use on CCW page.
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '//select[@name="intendedUse"]/optgroup[@label="All Intended Use"]/option[@kdfid="selectedPriclist_Internal Business Use"]'))).click()
    except Exception as e:
        if verbose:
            print('Failed during setting up value of Intended Use on CCW page. Error: {}'.format(e)) 
        else:
            print('Failed during setting up value of Intended Use on CCW page.')

    # Setting up value of Price List on CCW page.
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '//select[@name="priceList"]/optgroup[@label="All Price Lists"]/option[@kdfid="selectedPriclist_Global Price List EMEA Availability"]'))).click()
    except Exception as e:
        if verbose:
            print('Failed during setting up value of Price List on CCW page. Error: {}'.format(e)) 
        else:
            print('Failed during setting up value of Price List on CCW page.')

    # Setting up value of Buy Method on CCW page.
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '//select[@name="buyMethod"]/optgroup[@label="All Buy Method"]/option[@value="1"]'))).click()
    except Exception as e:
        if verbose:
            print('Failed during setting up value of Buy Method on CCW page. Error: {}'.format(e)) 
        else:
            print('Failed during setting up value of Buy Method on CCW page.')

    # Clicking Create Quote button on CCW page.
    # try:
    #     WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '//button[@kdfid="createQuote"]'))).click()
    # except Exception as e:
    #     if verbose:
    #         print('Failed during clicking Create Quote button on CCW page. Error: {}'.format(e)) 
    #     else:
    #         print('Failed during clicking Create Quote button on CCW page.')

    return


def create_new_cpq_quote(driver, opp, timeout_timer, verbose = True):
    
    # Selecting View All (quotes)
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '//span[@class="view-all-label"]'))).click()
    except Exception as e:
        if verbose:
            print('Failed when selecting View All (quotes). Error message: {}'.format(e.msg))
        else:
            print('Failed when selecting View All (quotes).')

    # Selecting CCW Quote Link
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, opp.quote_name))).click()
    except Exception as e:
        if verbose:
            print('Failed when selecting CPQ Quote Link. Error message: {}'.format(e.msg))
        else:
            print('Failed when selecting CPQ Quote Link.')

    # Navigating to CCW quote page
    # Store the ID of the original window
    saleforce_start_window = driver.current_window_handle
    # Wait for the new window or tab
    WebDriverWait(driver, timeout_timer).until(EC.number_of_windows_to_be(2))
    # Loop through until we find a new window handle
    for window_handle in driver.window_handles:
        if window_handle != saleforce_start_window:
            driver.switch_to.window(window_handle)
            break

    # Selecting Quote Tab on CCW quote page
    try:
        time.sleep(10)
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/nav/div/div/ul/li[2]/a/span[1]'))).click()
    except Exception as e:
        if verbose:
            print('Failed selecting Quote Tab on new CPQ quote Creation page for quote: {}. Error message: {}'.format(opp.quote_name, e.msg))
        else:
            print('Failed selecting Quote Tab on new CPQ quote Creation page for quote: {}.'.format(opp.quote_name))

    # Selecting Actions on Quote Tab on CCW quote page
    try:
        time.sleep(3)
        WebDriverWait(driver, timeout_timer).until(EC.invisibility_of_element((By.XPATH, '/html/body/div[1]/div[1]')))
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[5]/div/div[1]/div[4]/div[1]/div[1]/div[1]/div/div[2]/div/ul/li[3]/ul/li/span'))).click()
    except Exception as e:
        if verbose:
            print('Failed selecting Actions on Quote Tab on new CPQ quote Creation page for quote: {}. Error message: {}'.format(opp.quote_name, e.msg))
        else:
            print('Failed selecting Actions on Quote Tab on new CPQ quote Creation page for quote: {}.'.format(opp.quote_name))


    # Selecting Adding Advanced Services on Quote Tab on CCW quote page
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div[5]/div/div[1]/div[4]/div[1]/div[1]/div[1]/div/div[2]/div/ul/li[3]/ul/li/ul/li[8]/a'))).click()
    except Exception as e:
        if verbose:
            print('Failed selecting Adding Advanced Services on Quote Tab on new CPQ quote Creation page for quote: {}. Error message: {}'.format(opp.quote_name, e.msg))
        else:
            print('Failed selecting Adding Advanced Services on Quote Tab on new CPQ quote Creation page for quote: {}.'.format(opp.quote_name))

    # Selecting Buy Method on Quote Tab on CCW quote page
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div/div[2]/div/div[2]/div[2]/div/span/select')))
        dropdown = Select(driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/div[2]/div[2]/div/span/select'))
        dropdown.select_by_visible_text('Cisco')
    except Exception as e:
        if verbose:
            print('Failed selecting Buy Method on Quote Tab on new CPQ quote Creation page for quote: {}. Error message: {}'.format(opp.quote_name, e.msg))
        else:
            print('Failed selecting Buy Method on Quote Tab on new CPQ quote Creation page for quote: {}.'.format(opp.quote_name))

    # Confirming Buy Method on Quote Tab on CCW quote page
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div/div[2]/div/div[3]/div[2]/div[1]/div[2]/button[2]'))).click()
    except Exception as e:
        if verbose:
            print('Failed confirming Buy Method on Quote Tab on new CPQ quote Creation page for quote: {}. Error message: {}'.format(opp.quote_name, e.msg))
        else:
            print('Failed confirming Buy Method on Quote Tab on new CPQ quote Creation page for quote: {}.'.format(opp.quote_name))

    # Selecting Continue on new CPQ quote Creation page
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="brandBand_2"]/div/section/div/footer/button[2]'))).click()
    except Exception as e:
        if verbose:
            print('Failed selecting Continue on new CPQ quote Creation page for quote: {}. Error message: {}'.format(opp.quote_name, e.msg))
        else:
            print('Failed selecting Continue on new CPQ quote Creation page for quote: {}.'.format(opp.quote_name))


    # Entering Estimate Name on new CPQ quote Creation page
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section[1]/div/div/section/div/div[2]/div/div/section/div/div/div/div/div[2]/form/div/div/div/lightning-input/lightning-primitive-input-simple/div/div/input'))).send_keys(opp.quote_name)
    except Exception as e:
        if verbose:
            print('Failed Entering Estimate Name on new CPQ quote Creation page for quote: {}. Error message: {}'.format(opp.quote_name, e.msg))
        else:
            print('Failed Entering Estimate Name on new CPQ quote Creation page for quote: {}.'.format(opp.quote_name))

    # Selecting Offer Type on new CPQ quote Creation page
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section[1]/div/div/section/div/div[2]/div/div/section/div/div/div/div/div[5]/div[1]/div/div[1]/div/select')))
        dropdown = Select(driver.find_element(By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section[1]/div/div/section/div/div[2]/div/div/section/div/div/div/div/div[5]/div[1]/div/div[1]/div/select'))
        dropdown.select_by_visible_text('AST')
    except Exception as e:
        if verbose:
            print('Failed selecting Buy Method on Quote Tab on new CPQ quote Creation page for quote: {}. Error message: {}'.format(opp.quote_name, e.msg))
        else:
            print('Failed selecting Buy Method on Quote Tab on new CPQ quote Creation page for quote: {}.'.format(opp.quote_name))

    # Selecting Estimate Type on new CPQ quote Creation page
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section[1]/div/div/section/div/div[2]/div/div/section/div/div/div/div/div[5]/div[2]/div/div/div/select')))
        dropdown = Select(driver.find_element(By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section[1]/div/div/section/div/div[2]/div/div/section/div/div/div/div/div[5]/div[2]/div/div/div/select'))
        dropdown.select_by_visible_text('New Sale')
    except Exception as e:
        if verbose:
            print('Failed selecting Estimate Type on Quote Tab on new CPQ quote Creation page for quote: {}. Error message: {}'.format(opp.quote_name, e.msg))
        else:
            print('Failed selecting Estimate Type on Quote Tab on new CPQ quote Creation page for quote: {}.'.format(opp.quote_name))

    # Selecting Project Sub-Type on new CPQ quote Creation page
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section[1]/div/div/section/div/div[2]/div/div/section/div/div/div/div/div[5]/div[3]/div/div[1]/div/select')))
        dropdown = Select(driver.find_element(By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section[1]/div/div/section/div/div[2]/div/div/section/div/div/div/div/div[5]/div[3]/div/div[1]/div/select'))
        dropdown.select_by_visible_text('SOW [All Others]')
    except Exception as e:
        if verbose:
            print('Failed selecting Project Sub-Type on Quote Tab on new CPQ quote Creation page for quote: {}. Error message: {}'.format(opp.quote_name, e.msg))
        else:
            print('Failed selecting Project Sub-Type on Quote Tab on new CPQ quote Creation page for quote: {}.'.format(opp.quote_name))


# Selecting NGSC Theater on new CPQ quote Creation page
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section[1]/div/div/section/div/div[2]/div/div/section/div/div/div/div/div[6]/div[5]/div/div/div[1]/div/select')))
        dropdown = Select(driver.find_element(By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section[1]/div/div/section/div/div[2]/div/div/section/div/div/div/div/div[6]/div[5]/div/div/div[1]/div/select'))
        dropdown.select_by_value(opp.ngsc_theater)
    except Exception as e:
        if verbose:
            print('Failed selecting NGSC Theater on new CPQ quote Creation page for quote: {}. Error message: {}'.format(opp.quote_name, e.msg))
        else:
            print('Failed selecting NGSC Theater on Quote Tab on new CPQ quote Creation page for quote: {}.'.format(opp.quote_name))

# Selecting Region on new CPQ quote Creation page
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section[1]/div/div/section/div/div[2]/div/div/section/div/div/div/div/div[6]/div[4]/div/div/div/div/select')))
        dropdown = Select(driver.find_element(By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section[1]/div/div/section/div/div[2]/div/div/section/div/div/div/div/div[6]/div[4]/div/div/div/div/select'))
        dropdown.select_by_value(opp.delivery_region)
    except Exception as e:
        if verbose:
            print('Failed selecting Region on new CPQ quote Creation page for quote: {}. Error message: {}'.format(opp.quote_name, e.msg))
        else:
            print('Failed selecting Region on Quote Tab on new CPQ quote Creation page for quote: {}.'.format(opp.quote_name))

# Selecting NGSC Region on new CPQ quote Creation page
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section[1]/div/div/section/div/div[2]/div/div/section/div/div/div/div/div[6]/div[6]/div/div/div[1]/div/select')))
        dropdown = Select(driver.find_element(By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section[1]/div/div/section/div/div[2]/div/div/section/div/div/div/div/div[6]/div[6]/div/div/div[1]/div/select'))
        dropdown.select_by_value(opp.ngsc_region)
    except Exception as e:
        if verbose:
            print('Failed selecting NGSC Region on new CPQ quote Creation page for quote: {}. Error message: {}'.format(opp.quote_name, e.msg))
        else:
            print('Failed selecting NGSC Region on Quote Tab on new CPQ quote Creation page for quote: {}.'.format(opp.quote_name))

    # Entering Service Seller Name on new CPQ quote Creation page
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section[1]/div/div/section/div/div[2]/div/div/section/div/div/div/div/div[6]/div[3]/div/div/div/div/div[2]/input'))).send_keys(opp.sales_name)
        time.sleep(2)
        name_list = driver.find_elements(By.XPATH, '//*[@id="ASTNewSale"]/div[3]/div/div/ul/li[@role="presentation"]')
        name_list[0].click()
    except Exception as e:
        if verbose:
            print('Failed Entering Service Seller Name on new CPQ quote Creation page for quote: {}. Error message: {}'.format(opp.quote_name, e.msg))
        else:
            print('Failed Entering Service Seller Name on new CPQ quote Creation page for quote: {}.'.format(opp.quote_name))

    # Entering Estimate Reviewer Name on new CPQ quote Creation page
    try:
        # WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section[1]/div/div/section/div/div[2]/div/div/section/div/div/div/div/div[6]/div[2]/div/div/div/div/div[2]/input'))).click()
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section[1]/div/div/section/div/div[2]/div/div/section/div/div/div/div/div[6]/div[2]/div/div/div/div/div[2]/input'))).send_keys(opp.reviewer_name)
        time.sleep(2)
        name_list = driver.find_elements(By.XPATH, '//*[@id="ASTNewSale"]/div[2]/div/div/ul/li[@role="presentation"]')
        name_list[0].click()
    except Exception as e:
        if verbose:
            print('Failed Entering Estimate Reviewer Name on new CPQ quote Creation page for quote: {}. Error message: {}'.format(opp.quote_name, e.msg))
        else:
            print('Failed Entering Estimate Reviewer Name on new CPQ quote Creation page for quote: {}.'.format(opp.quote_name))

    # Selecting Save on new CPQ quote Creation page
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section[1]/div/div/section/div/div[2]/div/div/section/div/footer/button[2]'))).click()
    except Exception as e:
        if verbose:
            print('Failed selecting Continue on new CPQ quote Creation page for quote: {}. Error message: {}'.format(opp.quote_name, e.msg))
        else:
            print('Failed selecting Continue on new CPQ quote Creation page for quote: {}.'.format(opp.quote_name))
    return


def edit_cpq_estimates_lines(driver, opp, timeout_timer, verbose = True):

    # Selecting Edit CPQ Estimates Lines
    try:
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Edit CPQ Estimate Lines"]'))).click()
    except Exception as e:
        if verbose:
            print('Failed when selecting Edit CPQ Estimates Lines. Error message: {}'.format(e.msg))
        else:
            print('Failed when selecting Edit CPQ Estimates Lines.')

    
    input('Press [Enter] to continue')

    # Pressing Suggest Button

    # //sb-dialog[@id="modal"]
    # //button[text()="Suggest"]
    # //div[@id="content"]
    

    # driver.find_element(By.XPATH, "/html/body/div[4]/div[1]/section/div[1]/div/div[2]/div[2]/section/div/div/section[4]/div/div[2]/div/div/div/force-aloha-page/div/iframe")

    iframes = driver.find_elements(By.XPATH, '//iframe')
    for iframe in iframes:
        print(iframe)
    
    try:
        driver.switch_to.frame(iframes[0])
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Suggest"]'))).click()
    except Exception as e:
        if verbose:
            print('Failed when Pressing Suggest Button on iFrame[0]. Error message: {}'.format(e.msg))
        else:
            print('Failed when Pressing Suggest Button on iFrame[0].')

    try:
        driver.switch_to.frame(iframes[1])
        WebDriverWait(driver, timeout_timer).until(EC.element_to_be_clickable((By.XPATH, '//button[text()="Suggest"]'))).click()
    except Exception as e:
        if verbose:
            print('Failed when Pressing Suggest Button on iFrame[1]. Error message: {}'.format(e.msg))
        else:
            print('Failed when Pressing Suggest Button on iFrame[1].')


    input('Press [Enter] to continue')
    iframe = driver.find_element(By.CSS_SELECTOR, ".slds-template_iframe > iframe:nth-child(1)")
    driver.switch_to.frame(iframe)
    driver.find_element(By.XPATH,'//paper-button[@id="suggest"]').click()


    # driver.find_element(By.XPATH,'//html').click()
    # driver.find_element(By.XPATH,'//html').send_keys(Keys.RETURN)


    # try:
    #     driver.find_element(By.ID, 'suggest').click()
    # except Exception as e:
    #     print(e)
    
    # iframes = driver.find_elements(By.XPATH, '//iframe')
    # driver.switch_to.frame(iframes[0])
    
    # try:
    #     driver.find_element(By.XPATH, '//paper-button[text()="Suggest"]').click()
    # except Exception as e:
    #     print(e)
    
    # try:
    #     driver.find_element(By.XPATH, '/html/body/span[1]/sb-page-container//div/sb-product-lookup//sb-dialog/sb-i18n').click()
    # except Exception as e:
    #     print(e)

    # try:
    #     driver.find_element(By.XPATH, '//div/div[2]/header/i[@class="sf-icon-close"]').click()
    # except Exception as e:
    #     print(e)

    # driver.switch_to.active_element    
    # print_all_elements_from_driver(driver)


    # input('Press [Enter] to continue')

    



    
    return



if __name__ == '__main__':

    """
    Scrypt to automate CPQ quote creation.
    """

    input_file_name = 'input-cpq-quote.txt'
    verbose = True
    timeout_timer = 30

    opp = Opportunity()

    srcript_input = {
        'open_mode': 'manual',
        'username': '',
        'filename': '',
        'max_MS_number' : 5,
        'opportunuty' : opp
    }

    read_script_input_data_file(input_file_name, srcript_input)

    # ------------------------------------
    # START - Reading LOE file
    # ------------------------------------
    # opp = read_opp_data_from_estimates_file(srcript_input['filename'],opp)
    # tasks_list = read_task_data_from_estimates_file(srcript_input['filename'], int(srcript_input['max_MS_number']))
    # tasks_list = correct_task_data(tasks_list)
    # print_text('Basic Quote information and task import data:')
    # print(opp)
    # for e in tasks_list:
    #     print(e)
    # ---------------------------------------------------------------------------
    #  END - Reading LOE file
    # ---------------------------------------------------------------------------
    # ---------------------------------------------------------------------------
    # START - Browser selection
    # ---------------------------------------------------------------------------
    driver = setup_web_driver('Firefox')
    driver.implicitly_wait(2)
    # ---------------------------------------------------------------------------
    # END - Browser selection
    # ---------------------------------------------------------------------------
    # ---------------------------------------------------------------------------
    # START - Open SalesForce
    # ---------------------------------------------------------------------------
    # # Production Sales Force
    web_page = 'https://ciscosales.my.salesforce.com'
    # Sandbox Sales Force
    # web_page = 'https://ciscosales--ts1sm.sandbox.lightning.force.com/'
    if srcript_input['open_mode'].startswith('ma'):
        driver.get(web_page)
        print_text('Please open SF main page and press Enter')
        input('Press [Enter] to continue')
    else:
        driver.get(web_page)
        print_text('Trying to go thru SSO and open ASPT Quoter Main page.')
        passing_thru_SSO(driver, srcript_input['username'], timeout_timer)
        time.sleep(10)
    # ---------------------------------------------------------------------------
    # END - Open SalesForce
    # ---------------------------------------------------------------------------
    # ---------------------------------------------------------------------------
    # START - SalesForce Operation
    # ---------------------------------------------------------------------------

    create_new_quote_in_CCW(driver, opp, timeout_timer, verbose = True)

    print_text('Trying to navigate to opportunity tab for DID: {}'.format(opp.DID))
    go_to_opportunity_tab(driver, opp, timeout_timer, verbose)

    print_text('Trying to navigate to CCW quote name: "{}" for DID: {}'.format(opp.quote_name, opp.DID))
    # Storing ID of Existing Window
    saleforce_start_window = driver.current_window_handle
    create_new_cpq_quote(driver, opp, timeout_timer, verbose = True)

    input('Press [Enter] to continue')

    print_text('Trying to edit CPQ quote lines: "{}" for DID: {}'.format(opp.quote_name, opp.DID))
    edit_cpq_estimates_lines(driver, opp, timeout_timer, verbose = True)
    
    print_text('All tasks have been completed')
    input('Press [Enter] to continue')
    
    driver.quit()


    # 56861081 - dummy DID